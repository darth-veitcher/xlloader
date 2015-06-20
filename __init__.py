from pprint import pprint
from openpyxl import load_workbook
from openpyxl.cell import get_column_letter, column_index_from_string
from collections import OrderedDict

def sql_safe_string(string):
    safe_chars = [32,]  # space
    [safe_chars.append(i) for i in range(48, 58)]  # 0 to 9
    [safe_chars.append(i) for i in range(65, 91)]  # A to Z
    [safe_chars.append(i) for i in range(97, 122)]  # a to z

    if not string:
        return ""

    output = ""
    for c in string:
        if ord(c) in safe_chars:
            output = output + c.replace(" ", "_")
    return output

def sheet_to_dict(file_path, sheet_name=None, header_row=1, start_col='A'
                    , sql_safe=True, keep_order=True):
    """ Imports a sheet from specified Excel file and returns a list of Python
        dictionaries with k,v corresponding to header row and values.

        Params:
          - file_path :  absolute path to the file
          - sheet_name (optional):  the sheet to import (default is the first)
          - header_row (optional):  the position of the header row (default is 1)
          - start_col (optional) :  the first column of data (default is A)
          - sql_safe (optional) :   will convert keys to a sql_safe string, replacing spaces
                                    with underscores (default is True)
          - keep_order (optional) : uses an OrderedDict to keep the data structure (column ordering)
                                    consistent with the spreadsheet (default is True)
    """
    # Optimised loader, use the data_only flag to ensure we get the values,
    # otherwise it will return the formula string in the cell as opposed to
    # what that formula resolves to...
    wb = load_workbook(file_path, read_only=True, data_only=True)

    if sheet_name:
        ws = wb.get_sheet_by_name(sheet_name)
    else:
        ws = wb.active  # default to the first sheet in file

    # Get the used range
    max_row = ws.get_highest_row()
    max_column = ws.get_highest_column()

    # Create list of header cell values to use as keys for the dict
    try:
        header_range = "{start_col}{header_row}:{col}{header_row}".format(
            start_col = start_col,
            header_row = header_row,
            col = get_column_letter(max_column)
        )
    except TypeError(e):
        print(e)
        print("Did you set the Header row correctly?!")

    if sql_safe:
        headers = tuple([sql_safe_string(cell.value) for row in ws[header_range] for cell in row])
    else:
        headers = tuple([cell.value for row in ws[header_range] for cell in row])

    # Get the rows
    rows_range = "{start_col}{start_row}:{end_col}{end_row}".format(
        start_col = start_col,
        end_col = get_column_letter(max_column - column_index_from_string(start_col) + 1),
        start_row = header_row + 1,
        end_row = max_row
    )

    rows = tuple(ws[rows_range])

    # Create the list of dicts for the project information
    if keep_order:
        dicts = [OrderedDict(zip(headers, [cell.value for cell in row])) for row in tuple(ws[rows_range])]
    else:
        dicts = [dict(zip(headers, [cell.value for cell in row])) for row in tuple(ws[rows_range])]

    return dicts
